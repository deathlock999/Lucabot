import requests
import openpyxl
import time
from deep_translator import GoogleTranslator



api_token = "rDfh1kJ7U6WBbx5qQA3mdvEAhUoRz3Hr"

url = "https://gate.whapi.cloud/messages/poll"

group_id = "120363261013619385@g.us"

excel_file_path = "Modified_excel.xlsx"

wb = openpyxl.load_workbook(excel_file_path)

sheet = wb.active
answer = []
correct_ans_set = []

last_question_index = 1


def update_index_file():
  try:
    with open("file.txt", 'r+') as f:
      # Handle empty file case (initialize to 0)
      content = f.read()
      last_index_id = int(content.strip()) if content else 0

      # Increment and update the file
      last_index_id += 1
      f.seek(0)
      f.write(str(last_index_id))

      # Return the updated last index ID
      return last_index_id
  except FileNotFoundError:
    # Handle case where the file doesn't exist (create it with initial value 0)
    with open("file.txt", 'w') as f:
      f.write('0')
      return 0

def translate(to_translate):
    translated = GoogleTranslator(source='auto', target='si').translate(to_translate)
    return translated


def get_answer(last_question_index):
    for i in range(2,6):
        if (len(answer)<4):
            cell = sheet.cell(last_question_index+1, i)
            to_translate = cell.value
            answers = translate(to_translate)
            answer.append(answers)
    return answer

def get_question(last_question_index):
    question_cell = sheet.cell(last_question_index+1,1)
    to_translate = question_cell.value
    questions = translate(to_translate)
    return questions    

def get_correct_ans(last_question_index):
    correct_ans_cell = sheet.cell(last_question_index+1,6)
    to_translate = correct_ans_cell.value
    correct_ans = translate(to_translate)
    return correct_ans


def send_poll(questions,answer):
    payload = {
        "to": '120363261013619385@g.us',
        "options": answer,
        "title": questions,
        "count": 1,
        "ephemeral": 10
    }
    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "authorization": api_token
    }
    response = requests.post(url, json=payload, headers=headers)

    print(response.text)

def send_message(api_token,group_id,correct_ans_set):
    print(correct_ans_set)
    
    base_url = "https://gate.whapi.cloud/"
    endpoint = "messages/text"
    message = f"**නිවැරදි පිළිතුරු**\n1.{correct_ans_set[0]}\n1.{correct_ans_set[1]}"
    correct_ans_set.clear()
    headers = {
        "Authorization": f"Bearer {api_token}"
    }

    data = {
        "to": group_id,
        "body": message
    }

    try:
        response = requests.post(url=f"{base_url}{endpoint}", headers=headers, json=data)
        response.raise_for_status()  # Raise an exception for non-2xx status codes
        return response.json()  # Return the JSON response if successful
        print("Sent")
    except requests.exceptions.RequestException as error:
        print(f"Error sending message: {error}")
        return None  # Indicate an error


while (True):
    for z in range(1,5):
        last_question_index = update_index_file()
        get_question(last_question_index)
        get_answer(last_question_index)
        get_correct_ans(last_question_index)
        #question and answer send...
        ques = get_question(last_question_index)
        ans = get_answer(last_question_index)
        send_poll(ques, ans)
        #correct answer...
        correct_ans_set.append(get_correct_ans(last_question_index))
        answer.clear()
        time.sleep(900)
    send_message(api_token, group_id, correct_ans_set)
    time.sleep(900)

